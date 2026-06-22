from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import uvicorn
import os
import shutil
import io
import json
from sqlalchemy.orm import Session
from fastapi import Depends, BackgroundTasks
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import openpyxl
import threading
from sqlalchemy import text
import re
import fitz
from PIL import Image


try:
    from . import models, schemas, database, ai_service, config_service
except ImportError:
    import models, schemas, database, ai_service, config_service

# Run SQLite table creation
models.Base.metadata.create_all(bind=database.engine)

# Migration: Add is_archived column to documents if it doesn't exist yet
try:
    with database.engine.connect() as conn:
        conn.execute(text("ALTER TABLE documents ADD COLUMN is_archived BOOLEAN DEFAULT 0"))
        conn.commit()
except Exception as e:
    # Column already exists or table not created yet
    pass

# Migration: Add link_contract_id column to documents if it doesn't exist yet
try:
    with database.engine.connect() as conn:
        conn.execute(text("ALTER TABLE documents ADD COLUMN link_contract_id INTEGER"))
        conn.commit()
except Exception as e:
    pass

# Migration: Add province and county columns to documents if they don't exist yet
try:
    with database.engine.connect() as conn:
        conn.execute(text("ALTER TABLE documents ADD COLUMN province VARCHAR"))
        conn.execute(text("ALTER TABLE documents ADD COLUMN county VARCHAR"))
        conn.commit()
except Exception as e:
    pass

# Migration: Add progress_status column to documents if it doesn't exist yet
try:
    with database.engine.connect() as conn:
        conn.execute(text("ALTER TABLE documents ADD COLUMN progress_status VARCHAR DEFAULT '等待中'"))
        conn.commit()
except Exception as e:
    pass

# Migration: Add source column to documents if it doesn't exist yet
try:
    with database.engine.connect() as conn:
        conn.execute(text("ALTER TABLE documents ADD COLUMN source VARCHAR DEFAULT 'file_archive'"))
        conn.commit()
except Exception as e:
    pass




app = FastAPI(title="Document Management System API")

excel_lock = threading.Lock()
ledger_lock_state = {"status": "ok", "message": ""}

# Session management: keep track of documents uploaded in the current session
current_session_doc_ids = set()

# Initialize current_session_doc_ids with any existing unarchived documents
try:
    db = database.SessionLocal()
    unarchived_ids = [doc.id for doc in db.query(models.Document).filter(models.Document.is_archived == False).all()]
    current_session_doc_ids.update(unarchived_ids)
    db.close()
    print(f"Initialized current_session_doc_ids with {len(unarchived_ids)} unarchived documents.")
except Exception as e:
    print("Error initializing current_session_doc_ids:", e)


def cleanup_ghost_archived_documents(db: Session):
    try:
        ghosts = db.query(models.Document).filter(models.Document.is_archived == True).all()
        to_delete = []
        for doc in ghosts:
            if not doc.filepath or not os.path.exists(doc.filepath):
                to_delete.append(doc.id)
        if to_delete:
            print(f"Cleaning up {len(to_delete)} ghost archived records from DB: {to_delete}")
            db.query(models.Document).filter(models.Document.id.in_(to_delete)).delete(synchronize_session=False)
            db.commit()
    except Exception as e:
        print("Failed to clean up ghost archived records:", e)


def cleanup_unarchived_session_documents(db: Session):
    try:
        unarchived_docs = db.query(models.Document).filter(models.Document.is_archived == False).all()
        if unarchived_docs:
            print(f"Startup cleanup: Found {len(unarchived_docs)} unarchived temporary documents to delete.")
            for doc in unarchived_docs:
                if doc.filepath and os.path.exists(doc.filepath):
                    try:
                        os.remove(doc.filepath)
                    except Exception as e:
                        print(f"Failed to remove temporary file {doc.filepath}: {e}")
            
            doc_ids = [doc.id for doc in unarchived_docs]
            db.query(models.Document).filter(models.Document.id.in_(doc_ids)).delete(synchronize_session=False)
            db.commit()
            print("Startup cleanup: Unarchived database records deleted successfully.")
    except Exception as e:
        print("Failed to run startup cleanup for unarchived documents:", e)


# Configure CORS for frontend access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for dev, restrict in prod
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def add_no_cache_header(request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/api"):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

@app.post("/api/log-error")
def log_error(data: dict):
    print("FRONTEND ERROR:", data)
    os.makedirs("scratch", exist_ok=True)
    with open("scratch/frontend_error.log", "a", encoding="utf-8") as f:
        f.write(json.dumps(data, ensure_ascii=False) + "\n")
    return {"status": "ok"}

# Header list for Excel Ledger (V7.5 new 17 columns layout)
DEFAULT_NEW_HEADERS = [
    "签订日期", "合同编号", "矿名", "合同名称", "产品名称", "规格型号", 
    "数量", "单价", "金额", "合同状态", "盖章状态", "发货状态", 
    "收货状态", "验收状态", "开票状态", "回款状态", "备注"
]

def parse_product_details_to_columns(products_str: str):
    if not products_str:
        return "", "", "", ""
    names = []
    specs = []
    qtys = []
    prices = []
    # Split by semicolon or newline
    entries = re.split(r'[;；\n]', products_str)
    for entry in entries:
        entry = entry.strip()
        if not entry:
            continue
        name = entry
        spec = ""
        qty = ""
        price = ""
        # Check parenthesized format: Name(content)
        match = re.match(r'^([^(]+)\(([^)]+)\)$', entry)
        if match:
            name = match.group(1).strip()
            content = match.group(2).strip()
            # Split content by comma
            parts = [p.strip() for p in re.split(r'[,，]', content)]
            for part in parts:
                if part.startswith("规格:") or part.startswith("规格："):
                    spec = part[3:].strip()
                elif "*" in part or "x" in part or "X" in part or "=" in part:
                    # e.g., 4个*1680=6720
                    lhs = part.split("=")[0].strip()
                    sub_parts = re.split(r'[\*xX]', lhs)
                    if len(sub_parts) >= 2:
                        q_str = sub_parts[0].strip()
                        p_str = sub_parts[1].strip()
                        q_match = re.search(r'(\d+(?:\.\d+)?)', q_str)
                        if q_match:
                            qty = q_match.group(1)
                        p_match = re.search(r'(\d+(?:\.\d+)?)', p_str)
                        if p_match:
                            price = p_match.group(1)
                    else:
                        num_match = re.search(r'(\d+(?:\.\d+)?)', part)
                        if num_match:
                            qty = num_match.group(1)
                else:
                    num_match = re.search(r'(\d+(?:\.\d+)?)', part)
                    if num_match:
                        qty = num_match.group(1)
        else:
            # Non-parenthesized fallback, e.g. Name 1套 26000元
            numbers = re.findall(r'(\d+(?:\.\d+)?)\s*(?:[个件台套把只吨支双卷包箱张本瓶盒kgmt元]|$)', entry)
            numbers = [n for n in numbers if n]
            if len(numbers) >= 2:
                qty = numbers[0]
                price = numbers[1]
                idx = entry.find(numbers[0])
                if idx != -1:
                    name = entry[:idx].strip()
            elif len(numbers) == 1:
                qty = numbers[0]
                idx = entry.find(numbers[0])
                if idx != -1:
                    name = entry[:idx].strip()
                    
        def format_num_str(s):
            if not s:
                return ""
            try:
                val = float(s)
                if val.is_integer():
                    return str(int(val))
                return str(val)
            except:
                return s
                
        names.append(name)
        specs.append(spec)
        qtys.append(format_num_str(qty))
        prices.append(format_num_str(price))
        
    return "\n".join(names), "\n".join(specs), "\n".join(qtys), "\n".join(prices)

def get_unique_contracts(db: Session) -> List[models.Document]:
    # Query all archived contracts ordered by Document ID ascending
    contracts = db.query(models.Document).filter(
        models.Document.document_type.in_(get_contract_types()),
        models.Document.is_archived == True
    ).order_by(models.Document.id.asc()).all()
    
    # Group contracts by unique key (contract number or buyer + date)
    # The later uploads (with higher IDs) will overwrite previous ones, preserving only the latest version.
    grouped = {}
    for doc in contracts:
        try:
            data = json.loads(doc.extracted_data or "{}")
        except:
            data = {}
            
        c_no = data.get("合同/发票编号", "").strip()
        buyer = data.get("买受方/购买方", "").strip()
        date = data.get("签订/开票日期", "").strip()
        
        if c_no:
            key = f"no:{c_no}"
        else:
            key = f"buyer_date:{buyer}|{date}"
            
        grouped[key] = doc
        
    return list(grouped.values())

def make_file_normal(path: str):
    if os.name == 'nt' and os.path.exists(path):
        try:
            import ctypes
            # FILE_ATTRIBUTE_NORMAL = 0x80
            ctypes.windll.kernel32.SetFileAttributesW(path, 0x80)
        except Exception as e:
            print(f"Failed to set normal attribute on {path}: {e}")

def make_file_hidden(path: str):
    if os.name == 'nt' and os.path.exists(path):
        try:
            import ctypes
            # FILE_ATTRIBUTE_HIDDEN = 2
            ctypes.windll.kernel32.SetFileAttributesW(path, 2)
        except Exception as e:
            print(f"Failed to set hidden attribute on {path}: {e}")

def get_contract_types() -> List[str]:
    try:
        config = config_service.get_config()
        return config.get("contract_types", ["合同", "销售合同"])
    except Exception as e:
        print("Failed to load contract types:", e)
        return ["合同", "销售合同"]

def update_excel_ledger_file_core(contract_doc: models.Document, db: Session, raise_http: bool = False):
    global ledger_lock_state
    if not contract_doc.is_archived:
        return
        
    config = config_service.get_config()
    archive_dir = config["archive_dir"]
    os.makedirs(archive_dir, exist_ok=True)
    excel_path = os.path.join(archive_dir, "台账.xlsx")
    
    # Load or create workbook
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "文档台账"
            ws.append(DEFAULT_NEW_HEADERS)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "文档台账"
        ws.append(DEFAULT_NEW_HEADERS)
        
    # Read headers
    headers_in_sheet = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map = {str(h).strip(): idx for idx, h in enumerate(headers_in_sheet, 1) if h}
    
    # Check if old format (contains '序号' or lacks '发货状态' or has '当前有效合同')
    is_old_layout = "序号" in header_map or "发货状态" not in header_map or "当前有效合同" in header_map
    if is_old_layout:
        # Migrate sheet to new 17 columns layout
        ws.delete_rows(1, ws.max_row + 1)
        ws.append(DEFAULT_NEW_HEADERS)
        headers_in_sheet = DEFAULT_NEW_HEADERS
        header_map = {h: idx for idx, h in enumerate(DEFAULT_NEW_HEADERS, 1)}
        
        # When migrating, rewrite only the latest version of each contract currently in SQLite db!
        all_contracts = get_unique_contracts(db)
        for c_doc in all_contracts:
            write_contract_to_sheet(ws, c_doc, db)
    else:
        # Write/update only this contract (will naturally overwrite the existing row)
        write_contract_to_sheet(ws, contract_doc, db)
        
    if os.path.exists(excel_path):
        try:
            backup_path = excel_path.replace(".xlsx", "_backup.xlsx")
            make_file_normal(backup_path)
            shutil.copy2(excel_path, backup_path)
            make_file_hidden(backup_path)
        except Exception as e:
            print("Failed to create backup:", e)
            ledger_lock_state.update({
                "status": "locked", 
                "message": "本地台账备份失败，可能文件正被 Excel 或 WPS 占用，请关闭文件后再试。"
            })
            if raise_http:
                raise HTTPException(
                    status_code=400,
                    detail="本地台账备份失败，可能正被 Excel 或 WPS 占用，请先关闭该 Excel 文件后再试！"
                )
            return

    try:
        wb.save(excel_path)
        ledger_lock_state.update({"status": "ok", "message": ""})
    except PermissionError:
        ledger_lock_state.update({
            "status": "locked", 
            "message": "本地台账（台账.xlsx）正被 Excel 或 WPS 占用，无法写入更新。请先关闭 Excel 软件。"
        })
        if raise_http:
            raise HTTPException(
                status_code=400, 
                detail="本地台账（台账.xlsx）正被 Excel 或其他程序占用，请先关闭该 Excel 文件后再试！"
            )
        else:
            print("WARNING: Excel ledger (台账.xlsx) is locked. Cannot save updates.")

def check_disk_statuses(contract_filepath: str):
    has_delivery = False
    has_receipt = False
    has_invoice = False
    has_payment = False
    
    if not contract_filepath:
        return has_delivery, has_receipt, has_invoice, has_payment
        
    try:
        # Grandparent directory is the contract folder
        contract_dir = os.path.dirname(os.path.dirname(os.path.abspath(contract_filepath)))
        if os.path.exists(contract_dir) and os.path.isdir(contract_dir):
            for sub in os.listdir(contract_dir):
                sub_path = os.path.join(contract_dir, sub)
                if not os.path.isdir(sub_path):
                    continue
                # Check contents of subdirectory
                files = [f for f in os.listdir(sub_path) if os.path.isfile(os.path.join(sub_path, f))]
                if not files:
                    continue
                    
                sub_lower = sub.lower()
                if "02" in sub_lower or "发货" in sub_lower or "收货" in sub_lower or "送货" in sub_lower:
                    has_delivery = True
                    # Check if any file indicates receipt
                    for f in files:
                        f_lower = f.lower()
                        if "已收货" in f_lower or "已签收" in f_lower or "收货单" in f_lower or "签收单" in f_lower or "送货单" in f_lower:
                            has_receipt = True
                elif "03" in sub_lower or "发票" in sub_lower:
                    has_invoice = True
                elif "04" in sub_lower or "回款" in sub_lower or "付款" in sub_lower:
                    has_payment = True
    except Exception as e:
        print("Error checking disk statuses:", e)
        
    return has_delivery, has_receipt, has_invoice, has_payment

def write_contract_to_sheet(ws, contract_doc: models.Document, db: Session):
    headers_in_sheet = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map = {str(h).strip(): idx for idx, h in enumerate(headers_in_sheet, 1) if h}
    
    # Helper to map target header to its column index
    def get_col_index(target_names):
        for name in target_names:
            if name in header_map:
                return header_map[name]
        return None
        
    col_date = get_col_index(["签订日期", "签订/开票日期"])
    col_no = get_col_index(["合同编号", "合同/发票编号"])
    col_buyer = get_col_index(["矿名", "买受方/购买方", "买受方", "购买方"])
    col_name = get_col_index(["合同名称", "文档类型"])
    col_prod_name = get_col_index(["产品名称", "产品明细"])
    col_spec = get_col_index(["规格型号"])
    col_qty = get_col_index(["数量"])
    col_price = get_col_index(["单价"])
    col_amount = get_col_index(["金额", "价税合计金额"])
    col_contract_status = get_col_index(["合同状态", "状态"])
    col_seal_status = get_col_index(["盖章状态"])
    col_delivery_status = get_col_index(["发货状态"])
    col_receipt_status = get_col_index(["收货状态"])
    col_inspect_status = get_col_index(["验收状态"])
    col_billing_status = get_col_index(["开票状态"])
    col_payment_status = get_col_index(["回款状态"])
    col_comment = get_col_index(["备注"])
    
    # Parse extracted_data
    data_dict = {}
    if contract_doc.extracted_data:
        try:
            data_dict = json.loads(contract_doc.extracted_data)
        except:
            pass
            
    # Contract values
    contract_no = data_dict.get("合同/发票编号", "").strip()
    buyer = data_dict.get("买受方/购买方", "").strip()
    date = data_dict.get("签订/开票日期", "").strip()
    
    # Find all archived similar/duplicate contracts to aggregate linked documents
    same_contract_ids = [contract_doc.id]
    if contract_no:
        similar_contracts = db.query(models.Document).filter(
            models.Document.document_type.in_(get_contract_types()),
            models.Document.is_archived == True
        ).all()
        for c in similar_contracts:
            try:
                c_data = json.loads(c.extracted_data or "{}")
                if c_data.get("合同/发票编号", "").strip() == contract_no:
                    same_contract_ids.append(c.id)
            except:
                pass
    else:
        similar_contracts = db.query(models.Document).filter(
            models.Document.document_type.in_(get_contract_types()),
            models.Document.is_archived == True
        ).all()
        for c in similar_contracts:
            try:
                c_data = json.loads(c.extracted_data or "{}")
                c_buyer = c_data.get("买受方/购买方", "").strip()
                c_date = c_data.get("签订/开票日期", "").strip()
                if c_buyer == buyer and c_date == date:
                    same_contract_ids.append(c.id)
            except:
                pass

    # Query all archived linked documents linking to any duplicate contract record
    linked_docs = db.query(models.Document).filter(
        models.Document.link_contract_id.in_(same_contract_ids),
        models.Document.is_archived == True
    ).all()
    
    # Determine statuses
    has_delivery = False
    has_receipt = False
    has_invoice = False
    has_payment = False
    
    for d in linked_docs:
        d_type = get_mapped_doc_type(d.document_type)
        if d_type == "发票":
            has_invoice = True
        elif d_type == "回款":
            has_payment = True
        elif d_type == "收货单" or "收发货" in d.document_type or "发货" in d.document_type or "收货" in d.document_type or "送货" in d.document_type:
            d_data = {}
            if d.extracted_data:
                try:
                    d_data = json.loads(d.extracted_data)
                except:
                    pass
            if d_data.get("收货状态") == "已收货":
                has_receipt = True
                has_delivery = True
            else:
                has_delivery = True
                
    # Also check physical disk folders to avoid DB sync delay/missing records
    disk_delivery, disk_receipt, disk_invoice, disk_payment = check_disk_statuses(contract_doc.filepath)
    has_delivery = has_delivery or disk_delivery
    has_receipt = has_receipt or disk_receipt
    has_invoice = has_invoice or disk_invoice
    has_payment = has_payment or disk_payment
                
    # Seal status mapping
    raw_seal = data_dict.get("盖章状态", "").strip()
    if not raw_seal:
        seal_status = "未盖章/未识别"
    elif "双方盖章" in raw_seal or "双方已盖章" in raw_seal:
        seal_status = "双方已盖章"
    elif "单方盖章" in raw_seal or "单方已盖章" in raw_seal:
        seal_status = "单方已盖章"
    elif "未盖章" in raw_seal:
        seal_status = "未盖章/未识别"
    else:
        seal_status = raw_seal
        
    # Contract status mapping
    has_any_linked = (has_delivery or has_receipt or has_invoice or has_payment)
    if has_any_linked:
        contract_status = "履行中"
    elif seal_status == "双方已盖章":
        contract_status = "已生效"
    else:
        contract_status = seal_status
        
    # Other statuses
    delivery_status = "已发货" if has_delivery else ""
    receipt_status = "已签收/已收货" if has_receipt else ""
    inspect_status = "已验收" if has_receipt else ""
    billing_status = "已开票" if has_invoice else ""
    payment_status = "已回款" if has_payment else ""
    
    # Format amount
    amount_val = data_dict.get("价税合计金额") or data_dict.get("金额") or ""
    if amount_val:
        clean_amt = format_amount_clean(amount_val)
        amount_str = f"¥{clean_amt}"
    else:
        amount_str = ""
        
    # Search if contract row exists in Excel sheet
    target_row = None
    for r in range(2, ws.max_row + 1):
        # 1. Match by contract number
        if col_no and contract_no:
            val_no = str(ws.cell(row=r, column=col_no).value or "").strip()
            if val_no == contract_no:
                target_row = r
                break
        # 2. Match by buyer and date
        if col_buyer and col_date and buyer and date:
            val_buyer = str(ws.cell(row=r, column=col_buyer).value or "").strip()
            val_date = str(ws.cell(row=r, column=col_date).value or "").strip()
            if val_buyer == buyer and val_date == date:
                target_row = r
                break
                
    if not target_row:
        target_row = ws.max_row + 1
        
    # Write values cell by cell with centered alignment and color coding
    align_center = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_green = openpyxl.styles.PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_yellow = openpyxl.styles.PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_none = openpyxl.styles.PatternFill(fill_type=None)
    
    def write_cell(col_idx, value):
        if col_idx is not None:
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            cell.alignment = align_center
            
            # Apply color fill based on status value
            val_str = str(value).strip()
            if val_str in ["已回款", "双方已盖章", "已生效", "已发货", "已签收/已收货", "已验收", "已开票", "履行中"]:
                cell.fill = fill_green
            elif val_str in ["单方已盖章", "单方盖章"]:
                cell.fill = fill_yellow
            else:
                cell.fill = fill_none
            
    write_cell(col_date, date)
    write_cell(col_no, contract_no)
    write_cell(col_buyer, buyer)
    write_cell(col_name, get_mapped_doc_type(contract_doc.document_type))
    
    # Product columns split
    prod_str = data_dict.get("产品明细", "")
    if col_spec is not None:
        p_names, p_specs, p_qtys, p_prices = parse_product_details_to_columns(prod_str)
        write_cell(col_prod_name, p_names)
        write_cell(col_spec, p_specs)
        write_cell(col_qty, p_qtys)
        write_cell(col_price, p_prices)
    else:
        write_cell(col_prod_name, prod_str)
        
    write_cell(col_amount, amount_str)
    write_cell(col_contract_status, contract_status)
    write_cell(col_seal_status, seal_status)
    write_cell(col_delivery_status, delivery_status)
    write_cell(col_receipt_status, receipt_status)
    write_cell(col_inspect_status, inspect_status)
    write_cell(col_billing_status, billing_status)
    write_cell(col_payment_status, payment_status)
    write_cell(col_comment, data_dict.get("备注", ""))

def update_excel_ledger_file(doc: models.Document, raise_http: bool = False):
    db = database.SessionLocal()
    try:
        with excel_lock:
            if doc.document_type in get_contract_types():
                update_excel_ledger_file_core(doc, db, raise_http)
            elif doc.link_contract_id:
                contract_doc = db.query(models.Document).filter(models.Document.id == doc.link_contract_id).first()
                if contract_doc:
                    update_excel_ledger_file_core(contract_doc, db, raise_http)
    except Exception as e:
        print("Failed to update excel ledger:", e)
    finally:
        db.close()

def refresh_all_excel_records(db: Session, raise_http: bool = False):
    global ledger_lock_state
    config = config_service.get_config()
    archive_dir = config["archive_dir"]
    excel_path = os.path.join(archive_dir, "台账.xlsx")
    if not os.path.exists(excel_path):
        return
        
    excel_lock.acquire()
    try:
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Check if old format (contains '序号' or lacks '发货状态' or has '当前有效合同')
            headers_in_sheet = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            header_map = {str(h).strip(): idx for idx, h in enumerate(headers_in_sheet, 1) if h}
            is_old_layout = "序号" in header_map or "发货状态" not in header_map or "当前有效合同" in header_map
            if is_old_layout:
                ws.delete_rows(1, ws.max_row + 1)
                ws.append(DEFAULT_NEW_HEADERS)
                
            contracts = get_unique_contracts(db)
            
            for c in contracts:
                write_contract_to_sheet(ws, c, db)
                
            if os.path.exists(excel_path):
                try:
                    backup_path = excel_path.replace(".xlsx", "_backup.xlsx")
                    make_file_normal(backup_path)
                    shutil.copy2(excel_path, backup_path)
                    make_file_hidden(backup_path)
                except Exception as e:
                    print("Failed to create backup in refresh:", e)
                    ledger_lock_state.update({
                        "status": "locked", 
                        "message": "本地台账备份失败，可能文件正被 Excel 或 WPS 占用，请关闭文件后再试。"
                    })
                    if raise_http:
                        raise HTTPException(
                            status_code=400,
                            detail="本地台账备份失败，可能正被 Excel 或 WPS 占用，请先关闭该 Excel 文件后再试！"
                        )
                    return

            wb.save(excel_path)
            ledger_lock_state.update({"status": "ok", "message": ""})
        except PermissionError:
            ledger_lock_state.update({
                "status": "locked", 
                "message": "本地台账（台账.xlsx）正被 Excel 或 WPS 占用，无法写入更新。请先关闭 Excel 软件。"
            })
            if raise_http:
                raise HTTPException(
                    status_code=400, 
                    detail="本地台账（台账.xlsx）正被 Excel 或其他程序占用，请先关闭该 Excel 文件后再试！"
                )
            else:
                print("WARNING: Excel ledger (台账.xlsx) is locked. Cannot refresh updates.")
        except Exception as e:
            print("Failed to refresh all Excel records:", e)
    finally:
        excel_lock.release()

def remove_from_excel_ledger(doc_ids: List[int], raise_http: bool = False):
    global ledger_lock_state
    config = config_service.get_config()
    archive_dir = config["archive_dir"]
    excel_path = os.path.join(archive_dir, "台账.xlsx")
    if not os.path.exists(excel_path):
        return
        
    db = database.SessionLocal()
    target_filepaths = set()
    target_contract_nos = set()
    target_buyer_dates = set()
    try:
        # Fetch filepaths and contract numbers for the given doc_ids
        docs = db.query(models.Document).filter(models.Document.id.in_(doc_ids)).all()
        for doc in docs:
            if doc.filepath:
                target_filepaths.add(os.path.abspath(doc.filepath))
            if doc.extracted_data:
                try:
                    data_dict = json.loads(doc.extracted_data)
                    c_no = data_dict.get("合同/发票编号", "").strip()
                    if c_no:
                        target_contract_nos.add(c_no)
                    buyer = data_dict.get("买受方/购买方", "").strip()
                    date = data_dict.get("签订/开票日期", "").strip()
                    if buyer and date:
                        target_buyer_dates.add((buyer, date))
                except:
                    pass
    except Exception as e:
        print("Failed to fetch documents for Excel removal:", e)
    finally:
        db.close()
        
    excel_lock.acquire()
    try:
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Find column indices
            headers_in_sheet = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            header_map = {str(h).strip(): idx for idx, h in enumerate(headers_in_sheet, 1) if h}
            
            col_filepath = header_map.get("当前有效合同")
            col_no = header_map.get("合同编号") or header_map.get("合同/发票编号")
            col_buyer = header_map.get("矿名") or header_map.get("买受方/购买方") or header_map.get("买受方") or header_map.get("购买方")
            col_date = header_map.get("签订日期") or header_map.get("签订/开票日期")
            col_id_old = header_map.get("序号")  # backwards compatibility if old layout
            
            # Loop backwards to safely delete
            for r in range(ws.max_row, 1, -1):
                should_delete = False
                
                # Check old ID column
                if col_id_old:
                    if ws.cell(row=r, column=col_id_old).value in doc_ids:
                        should_delete = True
                        
                # Check filepath column
                if not should_delete and col_filepath:
                    val_path = ws.cell(row=r, column=col_filepath).value
                    if val_path and os.path.abspath(str(val_path)) in target_filepaths:
                        should_delete = True
                        
                # Check contract number column
                if not should_delete and col_no:
                    val_no = str(ws.cell(row=r, column=col_no).value or "").strip()
                    if val_no and val_no in target_contract_nos:
                        should_delete = True
                        
                # Check buyer and date
                if not should_delete and col_buyer and col_date:
                    val_buyer = str(ws.cell(row=r, column=col_buyer).value or "").strip()
                    val_date = str(ws.cell(row=r, column=col_date).value or "").strip()
                    if val_buyer and val_date and (val_buyer, val_date) in target_buyer_dates:
                        should_delete = True
                        
                if should_delete:
                    ws.delete_rows(r, 1)
                    
            if os.path.exists(excel_path):
                try:
                    backup_path = excel_path.replace(".xlsx", "_backup.xlsx")
                    make_file_normal(backup_path)
                    shutil.copy2(excel_path, backup_path)
                    make_file_hidden(backup_path)
                except Exception as e:
                    print("Failed to create backup in remove:", e)
                    ledger_lock_state.update({
                        "status": "locked", 
                        "message": "本地台账备份失败，可能文件正被 Excel 或 WPS 占用，请关闭文件后再试。"
                    })
                    if raise_http:
                        raise HTTPException(
                            status_code=400,
                            detail="本地台账备份失败，可能正被 Excel 或 WPS 占用，请先关闭该 Excel 文件后再试！"
                        )
                    return
    
            wb.save(excel_path)
            ledger_lock_state.update({"status": "ok", "message": ""})
        except PermissionError:
            ledger_lock_state.update({
                "status": "locked", 
                "message": "本地台账（台账.xlsx）正被 Excel 或 WPS 占用，无法写入更新。请先关闭 Excel 软件。"
            })
            if raise_http:
                raise HTTPException(
                    status_code=400, 
                    detail="本地台账（台账.xlsx）正被 Excel 或其他程序占用，请先关闭该 Excel 文件后再试！"
                )
            else:
                print("WARNING: Excel ledger (台账.xlsx) is locked. Cannot delete records.")
        except Exception as e:
            print("Failed to delete from excel ledger:", e)
    finally:
        excel_lock.release()

@app.get("/api/status")
def get_status():
    return {"status": "ok"}
@app.on_event("startup")
def startup_reprocess_pending():
    def startup_tasks():
        db = database.SessionLocal()
        try:
            # Clean up any ghost archived records (physically deleted files)
            cleanup_ghost_archived_documents(db)
            
            # Clean up unarchived temporary files left from previous crashed session
            cleanup_unarchived_session_documents(db)
            
            # Refresh the excel ledger on startup to sync physical directories and status updates
            refresh_all_excel_records(db)
            
            # Find any docs left in 'pending' status from previous crashed runs
            pending_docs = db.query(models.Document).filter(models.Document.status == "pending").all()
            for doc in pending_docs:
                if os.path.exists(doc.filepath):
                    # Reprocess in a separate thread
                    t = threading.Thread(
                        target=process_document_bg,
                        args=(doc.id, doc.filepath, doc.file_type),
                        daemon=True
                    )
                    t.start()
                else:
                    doc.status = "failed"
                    doc.summary = "文件已丢失，无法重试。"
                    db.commit()
        except Exception as e:
            print("Error reprocessing pending documents on startup:", e)
        finally:
            db.close()

    # Run everything in a background thread to avoid blocking the server startup
    threading.Thread(target=startup_tasks, daemon=True).start()

def parse_single_product(entry: str):
    entry = entry.strip()
    if not entry:
        return None
        
    name = entry
    qty = None
    
    match = re.search(r'^([^(]+)\(([^)]+)\)$', entry)
    if match:
        name = match.group(1).strip()
        content = match.group(2).strip()
        parts = [p.strip() for p in re.split(r'[,，]', content)]
        for part in parts:
            if part.startswith("规格:"):
                continue
            qty_sub = re.split(r'[\*=]', part)[0].strip()
            num_match = re.match(r'^(\d+(?:\.\d+)?)', qty_sub)
            if num_match:
                try:
                    qty = float(num_match.group(1))
                    break
                except ValueError:
                    pass
    else:
        num_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:[个件台套把只吨支双卷包箱张本瓶盒kgmt元]|$)', entry)
        if num_match:
            try:
                qty = float(num_match.group(1))
                name_candidate = entry.replace(num_match.group(0), "").strip()
                if name_candidate:
                    name = name_candidate
            except ValueError:
                pass
                
    name_normalized = re.sub(r'\s+', '', name).lower()
    return name_normalized, qty

def parse_products_and_quantities(products_str: str) -> dict:
    if not products_str:
        return {}
    entries = re.split(r'[;；\n]', products_str)
    result = {}
    for entry in entries:
        res = parse_single_product(entry)
        if res:
            name, qty = res
            if name:
                if name in result:
                    if qty is not None:
                        if result[name] is not None:
                            result[name] += qty
                        else:
                            result[name] = qty
                else:
                    result[name] = qty
    return result

def compare_products_and_quantities(prod_str1: str, prod_str2: str) -> bool:
    dict1 = parse_products_and_quantities(prod_str1)
    dict2 = parse_products_and_quantities(prod_str2)
    
    if not dict1 and not dict2:
        return True
    if not dict1 or not dict2:
        return False
        
    if set(dict1.keys()) != set(dict2.keys()):
        return False
        
    for name in dict1:
        q1 = dict1[name]
        q2 = dict2[name]
        if q1 is not None and q2 is not None:
            if abs(q1 - q2) > 1e-4:
                return False
    return True

def check_details_duplicate_core(doc_id: Optional[int], buyer: str, seller: str, amount: str, products: str, db: Session) -> Optional[str]:
    buyer = buyer.strip()
    seller = seller.strip()
    amount_clean = format_amount_clean(amount)
    products = products.strip()
    
    if not buyer:
        return None
        
    query = db.query(models.Document).filter(models.Document.status == "processed")
    if doc_id is not None:
        query = query.filter(models.Document.id != doc_id)
    other_docs = query.all()
    
    for other in other_docs:
        # If the file has been physically deleted from the disk, bypass it as a duplicate
        if other.filepath and not os.path.exists(other.filepath):
            continue
            
        try:
            other_data = json.loads(other.extracted_data or "{}")
        except:
            continue
            
        o_buyer = (
            other_data.get("买受方/购买方") or 
            other_data.get("买受方") or 
            other_data.get("购买方") or 
            other_data.get("买方") or 
            ""
        ).strip()
        o_seller = (
            other_data.get("出卖方/销售方") or 
            other_data.get("出卖方") or 
            other_data.get("销售方") or 
            other_data.get("卖方") or 
            ""
        ).strip()
        o_amount_clean = format_amount_clean(other_data.get("价税合计金额") or other_data.get("金额") or "")
        o_products = (other_data.get("产品明细") or "").strip()
        
        # Check buyer substring match (and not empty)
        buyer_match = False
        if buyer and o_buyer:
            buyer_match = buyer.lower() in o_buyer.lower() or o_buyer.lower() in buyer.lower()
        elif not buyer and not o_buyer:
            buyer_match = True
            
        # Check seller substring match
        seller_match = False
        if seller and o_seller:
            seller_match = seller.lower() in o_seller.lower() or o_seller.lower() in seller.lower()
        elif not seller and not o_seller:
            seller_match = True
            
        # Amount match:
        # If both are empty or 0, they match.
        # Otherwise they must be equal.
        amount_match = False
        is_amount_zero_1 = amount_clean in ["0", ""]
        is_amount_zero_2 = o_amount_clean in ["0", ""]
        if is_amount_zero_1 and is_amount_zero_2:
            amount_match = True
        elif not is_amount_zero_1 and not is_amount_zero_2:
            amount_match = (amount_clean == o_amount_clean)
        
        product_match = compare_products_and_quantities(products, o_products)
        
        if buyer_match and seller_match and amount_match and product_match:
            status_label = "已归档" if other.is_archived else "待处理"
            return f"疑似与已有单据重复 (匹配单据: {other.filename}, 状态: {status_label})"
            
    return None


def check_details_duplicate(doc_id: int, db: Session):
    doc = db.query(models.Document).filter(models.Document.id == doc_id).first()
    if not doc or not doc.extracted_data:
        return None
        
    try:
        data = json.loads(doc.extracted_data)
    except:
        return None
        
    buyer = (
        data.get("买受方/购买方") or 
        data.get("买受方") or 
        data.get("购买方") or 
        data.get("买方") or 
        ""
    )
    seller = (
        data.get("出卖方/销售方") or 
        data.get("出卖方") or 
        data.get("销售方") or 
        data.get("卖方") or 
        ""
    )
    amount = data.get("价税合计金额") or data.get("金额") or ""
    products = data.get("产品明细") or ""
    
    return check_details_duplicate_core(doc_id, buyer, seller, amount, products, db)

def process_document_bg(document_id: int, file_path: str, mime_type: str):
    db = database.SessionLocal()
    try:
        def update_progress(status_text: str):
            db_new = database.SessionLocal()
            try:
                doc = db_new.query(models.Document).filter(models.Document.id == document_id).first()
                if doc:
                    doc.progress_status = status_text
                    db_new.commit()
            except Exception as e:
                print("Error updating progress:", e)
            finally:
                db_new.close()

        update_progress("等待中...")
        ai_result = ai_service.analyze_document(file_path, mime_type, on_progress=update_progress)
        
        db_doc = db.query(models.Document).filter(models.Document.id == document_id).first()
        if db_doc:
            db_doc.document_type = ai_result.get("document_type", "未知")
            db_doc.extracted_data = ai_result.get("extracted_data")
            db_doc.summary = ai_result.get("summary")
            db_doc.status = "processed" if ai_result.get("document_type") != "处理失败" else "failed"
            
            # Pre-resolve province and county in background thread to optimize archiving performance
            if db_doc.status == "processed":
                try:
                    update_progress("自动匹配省份/县区中...")
                    extracted = json.loads(db_doc.extracted_data or "{}")
                    buyer_name = (
                        extracted.get("买受方/购买方") or 
                        extracted.get("买受方") or 
                        extracted.get("购买方") or 
                        extracted.get("买方") or 
                        ""
                    )
                    prov, county = resolve_province_county(file_path, buyer_name)
                    db_doc.province = prov
                    db_doc.county = county
                except Exception as e:
                    print("Error resolving province/county in background:", e)
                    
                try:
                    update_progress("自动关联匹配合同中...")
                    auto_link_documents(db)
                except Exception as e:
                    print("Error during auto_link_documents:", e)
                    
                try:
                    update_progress("检查数据重复中...")
                    warning = check_details_duplicate(db_doc.id, db)
                    if warning:
                        extracted = json.loads(db_doc.extracted_data or "{}")
                        extracted["duplicate_warning"] = warning
                        db_doc.extracted_data = json.dumps(extracted, ensure_ascii=False)
                except Exception as e:
                    print("Error checking details duplicate in background:", e)
                    
            update_progress("解析完成" if db_doc.status == "processed" else "解析失败")
            db.commit()
    except Exception as e:
        print("Error processing document in background:", e)
    finally:
        db.close()




@app.post("/api/upload", response_model=schemas.Document)
async def upload_file(file: UploadFile = File(...), source: str = "file_archive", db: Session = Depends(database.get_db)):
    config = config_service.get_config()
    archive_dir = config["archive_dir"]
    os.makedirs(archive_dir, exist_ok=True)
    
    file_location = os.path.join(archive_dir, file.filename)
    with open(file_location, "wb+") as file_object:
        shutil.copyfileobj(file.file, file_object)
    
    # Save to DB initially as pending
    db_doc = models.Document(
        filename=file.filename,
        filepath=file_location,
        file_type=file.content_type,
        status="pending",
        is_archived=False,
        source=source
    )

    db.add(db_doc)
    db.commit()
    db.refresh(db_doc)
    
    current_session_doc_ids.add(db_doc.id)
    
    # Run AI processing in background in a native Thread for true multi-threaded parallel execution
    t = threading.Thread(
        target=process_document_bg,
        args=(db_doc.id, file_location, file.content_type),
        daemon=True
    )
    t.start()
    
    return db_doc

@app.get("/api/documents", response_model=List[schemas.Document])
def get_documents(skip: int = 0, limit: int = 100, db: Session = Depends(database.get_db)):
    # Clean up any ghost archived records (physically deleted files)
    cleanup_ghost_archived_documents(db)
    
    global ledger_lock_state
    # If the ledger was previously locked, try to refresh it now. If it succeeds, it resets status to ok.
    if ledger_lock_state.get("status") == "locked":
        try:
            refresh_all_excel_records(db, raise_http=False)
        except Exception as e:
            print("Auto-retry refresh in get_documents failed:", e)

    # If session has no documents, return empty list immediately to avoid empty in_ query issues
    if not current_session_doc_ids:
        return []
    documents = db.query(models.Document).filter(
        models.Document.id.in_(current_session_doc_ids)
    ).order_by(models.Document.upload_time.desc()).offset(skip).limit(limit).all()
    return documents

@app.get("/api/documents/candidate-contracts")
def get_candidate_contracts(buyer: str = "", amount: str = "", date_str: str = "", products: str = "", db: Session = Depends(database.get_db)):
    # Clean receipt parameters
    r_buyer = (buyer or "").strip()
    r_amount_clean = format_amount_clean(amount) if amount else ""
    r_date_clean = format_date_yyyymmdd(date_str) if date_str else ""
    
    # Fetch all archived contracts ordered by upload time descending
    contracts = db.query(models.Document).filter(
        models.Document.is_archived == True,
        models.Document.document_type.in_(get_contract_types())
    ).order_by(models.Document.upload_time.desc()).all()
    
    seen_signatures = set()
    candidates = []
    for doc in contracts:
        try:
            extracted = json.loads(doc.extracted_data or "{}")
        except:
            extracted = {}
            
        c_buyer = (
            extracted.get("买受方/购买方") or 
            extracted.get("买受方") or 
            extracted.get("购买方") or 
            extracted.get("买方") or 
            ""
        ).strip()
        
        c_date = extracted.get("签订/开票日期") or extracted.get("签订日期") or ""
        c_amount = extracted.get("价税合计金额") or extracted.get("金额") or ""
        c_products = extracted.get("产品明细") or ""
        c_code = (extracted.get("合同/发票编号") or "").strip()
        
        c_amount_clean = format_amount_clean(c_amount)
        c_date_clean = format_date_yyyymmdd(c_date)
        
        # Deduplicate candidates by contract number or data signature
        if c_code:
            sig = f"code:{c_code}"
        else:
            sig = f"data:{c_buyer}|{c_amount_clean}|{c_date_clean}|{c_products.strip()}"
            
        if sig in seen_signatures:
            continue
        seen_signatures.add(sig)
        
        # Calculate amount match
        amount_match = False
        if r_amount_clean and c_amount_clean == r_amount_clean:
            amount_match = True
            
        # Calculate date difference if both dates are valid
        date_diff_days = None
        if r_date_clean and c_date_clean and len(r_date_clean) == 8 and len(c_date_clean) == 8:
            try:
                from datetime import datetime
                d1 = datetime.strptime(r_date_clean, "%Y%m%d")
                d2 = datetime.strptime(c_date_clean, "%Y%m%d")
                date_diff_days = abs((d1 - d2).days)
            except:
                pass
                
        # Calculate buyer match score
        buyer_match = False
        if r_buyer and c_buyer:
            buyer_match = (r_buyer == c_buyer) or (r_buyer in c_buyer) or (c_buyer in r_buyer)
            
        # Calculate product match score
        product_match = False
        product_overlap_count = 0
        if products and doc.extracted_data:
            def extract_product_keywords(p_str):
                if not p_str:
                    return set()
                keywords = []
                for word in re.findall(r'[\u4e00-\u9fa5A-Za-z0-9\-]+', p_str):
                    if len(word) >= 2 and word not in ["规格", "单位", "数量", "个", "件", "台", "套", "把", "只", "吨", "支", "备注", "带程序"]:
                        keywords.append(word)
                return set(keywords)
            
            r_kws = extract_product_keywords(products)
            c_kws = extract_product_keywords(c_products)
            overlap = r_kws & c_kws
            product_overlap_count = len(overlap)
            if product_overlap_count > 0:
                product_match = True
                    
        # Calculate comprehensive score

        score = 0
        if buyer_match:
            score += 100
        if amount_match:
            score += 50
        if product_match:
            score += 30 + product_overlap_count * 5
            
        if date_diff_days is not None:
            score -= min(date_diff_days // 10, 20)  # Penalize large date difference
        
        candidates.append({
            "id": doc.id,
            "filename": doc.filename,
            "filepath": doc.filepath,
            "buyer": c_buyer,
            "date": c_date,
            "amount": c_amount,
            "date_diff_days": date_diff_days,
            "buyer_match": buyer_match,
            "amount_match": amount_match,
            "product_match": product_match,
            "score": score
        })
        
    # Sort candidates by comprehensive score (descending)
    candidates.sort(key=lambda x: x["score"], reverse=True)
    
    # Recommendation logic:
    # Recommend the top candidate if it matches the buyer, and matches either amount OR products, and dates are close (<= 30 days)
    recommended_id = None
    if candidates:
        top = candidates[0]
        if top["buyer_match"] and (top["amount_match"] or top["product_match"]) and top["date_diff_days"] is not None and top["date_diff_days"] <= 30:
            recommended_id = top["id"]
            
    return {
        "candidates": candidates,
        "recommended_id": recommended_id
    }


@app.get("/api/documents/archived", response_model=List[schemas.Document])
def get_archived_documents(
    buyer: Optional[str] = None,
    doc_type: Optional[str] = None,
    date_start: Optional[str] = None,
    date_end: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    # Clean up any ghost archived records (physically deleted files)
    cleanup_ghost_archived_documents(db)
    
    query = db.query(models.Document).filter(models.Document.is_archived == True)
    docs = query.order_by(models.Document.upload_time.desc()).all()
    
    filtered = []
    for doc in docs:
        extracted = {}
        if doc.extracted_data:
            try:
                extracted = json.loads(doc.extracted_data)
            except:
                pass
        
        c_buyer = (
            extracted.get("买受方/购买方") or 
            extracted.get("买受方") or 
            extracted.get("购买方") or 
            extracted.get("买方") or 
            ""
        ).strip()
        
        if buyer and buyer.strip():
            if buyer.strip().lower() not in c_buyer.lower() and buyer.strip().lower() not in doc.filename.lower():
                continue
                
        if doc_type and doc_type != "全部":
            if doc.document_type != doc_type and get_mapped_doc_type(doc.document_type) != get_mapped_doc_type(doc_type):
                continue
                
        c_date = extracted.get("签订/开票日期") or extracted.get("签订日期") or ""
        c_date_clean = format_date_yyyymmdd(c_date)
        
        if date_start and len(date_start) == 8:
            if not c_date_clean or c_date_clean < date_start:
                continue
        if date_end and len(date_end) == 8:
            if not c_date_clean or c_date_clean > date_end:
                continue
                
        filtered.append(doc)
        
    return filtered

@app.get("/api/documents/check-duplicate")
def check_duplicate_filename(filename: str, db: Session = Depends(database.get_db)):
    exists = db.query(models.Document).filter(models.Document.filename == filename).first() is not None
    return {"exists": exists}


@app.get("/api/documents/{document_id}", response_model=schemas.Document)
def get_document(document_id: int, db: Session = Depends(database.get_db)):
    db_doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if db_doc is None:
        raise HTTPException(status_code=404, detail="Document not found")
    return db_doc


@app.get("/api/documents/{document_id}/file")
def get_document_file(document_id: int, db: Session = Depends(database.get_db)):
    db_doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if db_doc is None or not db_doc.filepath:
        raise HTTPException(status_code=404, detail="Document file not found")
    if not os.path.exists(db_doc.filepath):
        raise HTTPException(status_code=404, detail="Physical file does not exist")
    return FileResponse(db_doc.filepath)



@app.delete("/api/documents/{document_id}")
def delete_document(document_id: int, db: Session = Depends(database.get_db)):
    db_doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if db_doc is None:
        raise HTTPException(status_code=404, detail="Document not found")
    
    link_contract_id = db_doc.link_contract_id
    doc_type = db_doc.document_type
    doc_id = db_doc.id

    # Delete temporary file if not archived yet
    if not db_doc.is_archived and db_doc.filepath:
        if os.path.exists(db_doc.filepath):
            try:
                os.remove(db_doc.filepath)
            except Exception as e:
                print(f"Failed to delete temporary file {db_doc.filepath}: {e}")
                
    db.delete(db_doc)
    db.commit()
    if document_id in current_session_doc_ids:
        current_session_doc_ids.remove(document_id)
        
    # After deleting, update the Excel ledger!
    if doc_type in get_contract_types():
        remove_from_excel_ledger([doc_id])
    elif link_contract_id:
        contract_doc = db.query(models.Document).filter(models.Document.id == link_contract_id).first()
        if contract_doc:
            update_excel_ledger_file(contract_doc)
            
    return {"message": "Document deleted successfully"}

class BulkDeleteSchema(BaseModel):
    ids: List[int]

@app.post("/api/documents/bulk-delete")
def bulk_delete_documents(data: BulkDeleteSchema, db: Session = Depends(database.get_db)):
    db_docs = db.query(models.Document).filter(models.Document.id.in_(data.ids)).all()
    contracts_to_remove = []
    linked_contract_ids = set()
    
    for db_doc in db_docs:
        # Gather ledger info before deleting
        if db_doc.document_type in get_contract_types():
            contracts_to_remove.append(db_doc.id)
        elif db_doc.link_contract_id:
            linked_contract_ids.add(db_doc.link_contract_id)
            
        # Delete temporary file if not archived yet
        if not db_doc.is_archived and db_doc.filepath:
            if os.path.exists(db_doc.filepath):
                try:
                    os.remove(db_doc.filepath)
                except Exception as e:
                    print(f"Failed to delete temporary file {db_doc.filepath}: {e}")
        db.delete(db_doc)
        if db_doc.id in current_session_doc_ids:
            current_session_doc_ids.remove(db_doc.id)
        
    db.commit()
    
    # Update Excel ledger
    if contracts_to_remove:
        remove_from_excel_ledger(contracts_to_remove)
    for c_id in linked_contract_ids:
        contract_doc = db.query(models.Document).filter(models.Document.id == c_id).first()
        if contract_doc:
            update_excel_ledger_file(contract_doc)
            
    return {"message": f"{len(db_docs)} documents deleted successfully"}


# Settings APIs
@app.get("/api/settings")
def get_settings():
    return config_service.get_config()

@app.get("/api/settings/ledger-status")
def get_ledger_status(db: Session = Depends(database.get_db)):
    global ledger_lock_state
    if ledger_lock_state.get("status") == "locked":
        try:
            refresh_all_excel_records(db, raise_http=False)
        except Exception as e:
            print("Auto-retry refresh in get_ledger_status failed:", e)
    return ledger_lock_state

class SettingsUpdateSchema(BaseModel):
    archive_dir: str

class DocTypesUpdateSchema(BaseModel):
    document_types: List[str]
    contract_types: List[str]

@app.post("/api/settings/archive-dir")
def set_archive_dir(data: SettingsUpdateSchema):
    if not data.archive_dir:
        raise HTTPException(status_code=400, detail="Path cannot be empty")
    return config_service.update_config({"archive_dir": data.archive_dir})

class InvoiceSettingsUpdateSchema(BaseModel):
    invoice_archive_dir: str

@app.post("/api/settings/invoice-archive-dir")
def set_invoice_archive_dir(data: InvoiceSettingsUpdateSchema):
    if not data.invoice_archive_dir:
        raise HTTPException(status_code=400, detail="Path cannot be empty")
    return config_service.update_config({"invoice_archive_dir": data.invoice_archive_dir})

@app.get("/api/settings/choose-invoice-dir")
def choose_invoice_directory():
    import webview
    window = webview.active_window()
    if not window:
        return {"status": "error", "message": "No active GUI window found"}
    
    selected = window.create_file_dialog(webview.FOLDER_DIALOG)
    if selected and len(selected) > 0:
        folder_path = selected[0]
        config_service.update_config({"invoice_archive_dir": folder_path})
        return {"status": "success", "invoice_archive_dir": folder_path}
    return {"status": "cancelled"}

@app.get("/api/settings/open-invoice-folder")
def open_invoice_folder():
    config = config_service.get_config()
    invoice_dir = config["invoice_archive_dir"]
    if not os.path.exists(invoice_dir):
        try:
            os.makedirs(invoice_dir, exist_ok=True)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to create directory: {e}")
    try:
        os.startfile(invoice_dir)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/settings/invoice-subdirs")
def get_invoice_subdirs():
    config = config_service.get_config()
    invoice_dir = config.get("invoice_archive_dir")
    if not invoice_dir or not os.path.exists(invoice_dir):
        return []
    try:
        subdirs = []
        for name in os.listdir(invoice_dir):
            full_path = os.path.join(invoice_dir, name)
            if os.path.isdir(full_path) and not name.startswith('.'):
                subdirs.append(name)
        return subdirs
    except Exception as e:
        print("Failed to list invoice subdirs:", e)
        return []

class CreateInvoiceSubdirSchema(BaseModel):
    name: str

@app.post("/api/settings/invoice-subdirs")
def create_invoice_subdir(data: CreateInvoiceSubdirSchema):
    config = config_service.get_config()
    invoice_dir = config.get("invoice_archive_dir")
    if not invoice_dir or not os.path.exists(invoice_dir):
        raise HTTPException(status_code=400, detail="发票归档根目录未配置或不存在，请先在设置中选择")
    name_clean = re.sub(r'[\\/*?:"<>|]', "", data.name).strip()
    if not name_clean:
        raise HTTPException(status_code=400, detail="无效的目录名称")
    target_path = os.path.join(invoice_dir, name_clean)
    try:
        os.makedirs(target_path, exist_ok=True)
        return {"status": "success", "name": name_clean}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建目录失败: {str(e)}")

@app.post("/api/settings/document-types")
def set_document_types(data: DocTypesUpdateSchema):
    return config_service.update_config({
        "document_types": data.document_types,
        "contract_types": data.contract_types
    })

@app.get("/api/settings/choose-dir")
def choose_directory():
    import webview
    window = webview.active_window()
    if not window:
        return {"status": "error", "message": "No active GUI window found"}
    
    selected = window.create_file_dialog(webview.FOLDER_DIALOG)
    if selected and len(selected) > 0:
        folder_path = selected[0]
        config_service.update_config({"archive_dir": folder_path})
        return {"status": "success", "archive_dir": folder_path}
    return {"status": "cancelled"}

@app.get("/api/settings/open-folder")
def open_archive_folder():
    # Sync Excel status columns in real-time by checking database + physical subfolders before opening
    db = database.SessionLocal()
    try:
        refresh_all_excel_records(db, raise_http=True)
    except HTTPException as he:
        raise he
    except Exception as e:
        print("Failed to sync ledger statuses on folder open:", e)
    finally:
        db.close()
        
    config = config_service.get_config()
    archive_dir = config["archive_dir"]
    if os.path.exists(archive_dir):
        try:
            os.startfile(archive_dir)
            return {"status": "success"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    else:
        raise HTTPException(status_code=404, detail="Archive directory not found")

def parse_address_from_text(full_text: str, buyer_name: str = "") -> tuple:
    prov, city, county = "", "", ""
    pattern = r'(?:([\u4e00-\u9fa5]{2,10}(?:省|自治区|特别行政区))|北京市|上海市|天津市|重庆市)\s*(?:([\u4e00-\u9fa5]{2,10}(?:市|自治州|地区)))?\s*([\u4e00-\u9fa5]{2,10}(?:县|区|市|旗|盟))'
    
    # 1. Search lines with indicator keywords
    lines = full_text.split('\n')
    for line in lines:
        if any(keyword in line for keyword in ["地址", "住所", "地点", "住所地", "注册地", "买受方", "买方", "签订"]):
            match = re.search(pattern, line)
            if match:
                prov = match.group(1) or ""
                city = match.group(2) or ""
                county = match.group(3) or ""
                break
                
    # 2. If not found, search the whole text
    if not prov or not county:
        match = re.search(pattern, full_text)
        if match:
            prov = prov or match.group(1) or ""
            city = city or match.group(2) or ""
            county = county or match.group(3) or ""
            
    # 3. Special handling for municipalities
    for m in ["北京市", "上海市", "天津市", "重庆市"]:
        if m in full_text and not prov:
            prov = m
            
    # 4. Guess from buyer_name if province still missing
    if not prov and buyer_name:
        for p in ["贵州", "四川", "云南", "重庆", "北京", "上海", "天津", "广东", "浙江", "江苏", "山东", "河南", "河北", "山西", "陕西", "甘肃", "青海", "吉林", "辽宁", "黑龙江", "江西", "湖南", "湖北", "安徽", "福建", "海南", "内蒙古", "新疆", "西藏", "宁夏", "广西"]:
            if p in buyer_name:
                prov = p + ("市" if p in ["北京", "上海", "天津", "重庆"] else "省")
                break
                
    # 5. Clean up names
    if prov:
        prov = prov.strip()
    if county:
        county = county.strip()
    elif city:
        county = city.strip()
        
    return prov, county

def resolve_province_county(file_path: str, buyer_name: str) -> tuple:
    prov, county = "", ""
    try:
        full_text = ""
        if file_path.lower().endswith(".pdf"):
            with fitz.open(file_path) as doc:
                for page in doc:
                    full_text += page.get_text() + "\n"
            if len(full_text.strip()) < 50:
                full_text = ai_service.extract_text_from_pdf(file_path)
        elif file_path.lower().endswith((".jpg", ".jpeg", ".png", ".bmp", ".webp")):
            with Image.open(file_path) as img:
                full_text = ai_service.extract_text_from_image(img)
        else:
            # Fallback for plain text files or other text-readable files
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    full_text = f.read()
            except:
                pass
            
        prov, county = parse_address_from_text(full_text, buyer_name)
    except Exception as e:
        print(f"Error in resolve_province_county: {e}")
        
    if not prov:
        prov = "未知省份"
    if not county:
        county = "未知地区"
        
    return prov, county


def format_date_yyyymmdd(date_str: str) -> str:
    if not date_str:
        return "未知日期"
    # Find all digits in the string
    digits = re.sub(r'\D', '', date_str)
    if len(digits) == 8:
        return digits
    # If not 8 digits, try cleaning up common formats like YYYY-MM-DD, YYYY.MM.DD
    match = re.search(r'(\d{4})[-./年\s]*(\d{1,2})[-./月\s]*(\d{1,2})', date_str)
    if match:
        year = match.group(1)
        month = match.group(2).zfill(2)
        day = match.group(3).zfill(2)
        return f"{year}{month}{day}"
    return re.sub(r'[\\/*?:"<>|]', "", date_str).strip() or "未知日期"


def format_amount_clean(amount_str: str) -> str:
    if not amount_str:
        return "0"
    cleaned = amount_str.replace("¥", "").replace("￥", "").replace(",", "").strip()
    match = re.search(r'[-+]?\d*\.?\d+', cleaned)
    if not match:
        return "0"
    num_str = match.group(0)
    try:
        val = float(num_str)
        if val.is_integer():
            return str(int(val))
        else:
            s = f"{val:.2f}"
            if s.endswith(".00"):
                return s[:-3]
            elif s.endswith("0") and "." in s:
                return s[:-1]
            return s
    except ValueError:
        return re.sub(r'[\\/*?:"<>|]', "", amount_str).strip() or "0"


def get_mapped_doc_type(doc_type: str) -> str:
    mapping = {
        "合同": "销售合同",
        "收发货单": "收货单",
        "发票": "发票",
        "回款凭证": "回款"
    }
    return mapping.get(doc_type, doc_type or "其他")


def get_mapped_seal_status(seal_status: str) -> str:
    if not seal_status:
        return "无法确认"
    if "双方盖章" in seal_status:
        return "双方已盖章"
    if "单方盖章" in seal_status:
        return "单方已盖章"
    if "未盖章" in seal_status:
        return "未盖章"
    if "无法确认" in seal_status:
        return "无法确认"
    return "无法确认"



def auto_link_documents(db: Session):
    # Find all processed non-contract documents
    non_contracts = db.query(models.Document).filter(
        models.Document.status == "processed",
        models.Document.document_type.notin_(get_contract_types())
    ).all()
    
    # Find all contracts (both archived and unarchived)
    contracts = db.query(models.Document).filter(
        models.Document.status == "processed",
        models.Document.document_type.in_(get_contract_types())
    ).all()
    
    if not non_contracts or not contracts:
        return
        
    for doc in non_contracts:
        # Get extracted data for doc
        try:
            doc_data = json.loads(doc.extracted_data or "{}")
        except:
            continue
            
        buyer = (
            doc_data.get("买受方/购买方") or 
            doc_data.get("买受方") or 
            doc_data.get("购买方") or 
            doc_data.get("买方") or 
            ""
        ).strip()
        
        amount = doc_data.get("价税合计金额") or doc_data.get("金额") or ""
        date_str = doc_data.get("签订/开票日期") or doc_data.get("签订日期") or doc_data.get("开票日期") or doc_data.get("日期") or ""
        products = doc_data.get("产品明细") or ""
        
        r_buyer = buyer
        r_amount_clean = format_amount_clean(amount) if amount else ""
        r_date_clean = format_date_yyyymmdd(date_str) if date_str else ""
        
        best_contract_id = None
        best_score = 0
        
        for c in contracts:
            try:
                c_data = json.loads(c.extracted_data or "{}")
            except:
                continue
                
            c_buyer = (
                c_data.get("买受方/购买方") or 
                c_data.get("买受方") or 
                c_data.get("购买方") or 
                c_data.get("买方") or 
                ""
            ).strip()
            
            c_date = c_data.get("签订/开票日期") or c_data.get("签订日期") or ""
            c_amount = c_data.get("价税合计金额") or c_data.get("金额") or ""
            
            c_amount_clean = format_amount_clean(c_amount)
            c_date_clean = format_date_yyyymmdd(c_date)
            
            # 1. Buyer match
            buyer_match = False
            if r_buyer and c_buyer:
                buyer_match = (r_buyer == c_buyer) or (r_buyer in c_buyer) or (c_buyer in r_buyer)
                
            # 2. Amount match
            amount_match = False
            if r_amount_clean and c_amount_clean == r_amount_clean:
                amount_match = True
                
            # 3. Product match
            product_match = False
            product_overlap_count = 0
            if products and c_data.get("产品明细"):
                c_products = c_data.get("产品明细") or ""
                def extract_product_keywords(p_str):
                    if not p_str:
                        return set()
                    keywords = []
                    for word in re.findall(r'[\u4e00-\u9fa5A-Za-z0-9\-]+', p_str):
                        if len(word) >= 2 and word not in ["规格", "单位", "数量", "个", "件", "台", "套", "把", "只", "吨", "支", "备注", "带程序"]:
                            keywords.append(word)
                    return set(keywords)
                
                r_kws = extract_product_keywords(products)
                c_kws = extract_product_keywords(c_products)
                overlap = r_kws & c_kws
                product_overlap_count = len(overlap)
                if product_overlap_count > 0:
                    product_match = True
                    
            # 4. Date diff
            date_diff_days = None
            if r_date_clean and c_date_clean and len(r_date_clean) == 8 and len(c_date_clean) == 8:
                try:
                    from datetime import datetime
                    d1 = datetime.strptime(r_date_clean, "%Y%m%d")
                    d2 = datetime.strptime(c_date_clean, "%Y%m%d")
                    date_diff_days = abs((d1 - d2).days)
                except:
                    pass
                    
            # Check score
            score = 0
            if buyer_match:
                score += 100
            if amount_match:
                score += 50
            if product_match:
                score += 30 + product_overlap_count * 5
            if date_diff_days is not None:
                score -= min(date_diff_days // 10, 20)
                
            # Must match buyer, and either amount or products, and date difference must be <= 30 days
            is_good_match = buyer_match and (amount_match or product_match) and (date_diff_days is not None and date_diff_days <= 30)
            
            if is_good_match and score > best_score:
                best_score = score
                best_contract_id = c.id
                
        if best_contract_id:
            doc.link_contract_id = best_contract_id
            db.commit()


# Manual Archiving API (with data review edits)

def sync_invoice_history_with_db(invoice_archive_dir: str, db: Session):
    history_file_path = os.path.join(invoice_archive_dir, ".invoice_history.json")
    
    # 1. Load existing records
    records = []
    if os.path.exists(history_file_path):
        try:
            with open(history_file_path, "r", encoding="utf-8") as f:
                records = json.load(f)
        except Exception as e:
            print(f"Failed to read invoice history: {e}")
            
    # 2. Get all archived invoices from DB
    db_invoices = db.query(models.Document).filter(
        models.Document.document_type == "发票",
        models.Document.is_archived == True,
        models.Document.source == "invoice_archive"
    ).all()
    
    # Create a map of existing records by invoice_no/filename to avoid duplicates
    existing_keys = set()
    for r in records:
        # Check physical existence of record
        record_file_path = os.path.join(invoice_archive_dir, r.get("personal_dir", ""), r.get("filename", ""))
        if os.path.exists(record_file_path):
            existing_keys.add((r.get("invoice_no"), r.get("filename"), r.get("personal_dir")))
            
    # Rebuild records list using only physically existing files
    synced_records = []
    for r in records:
        record_file_path = os.path.join(invoice_archive_dir, r.get("personal_dir", ""), r.get("filename", ""))
        if os.path.exists(record_file_path):
            synced_records.append(r)
            
    # 3. Add any missing archived invoices from DB that physically exist
    has_new = False
    for doc in db_invoices:
        if not doc.filepath or not os.path.exists(doc.filepath):
            continue
            
        try:
            extracted = json.loads(doc.extracted_data or "{}")
        except:
            continue
            
        invoice_no = (
            extracted.get("合同/发票编号") or
            extracted.get("发票号码") or
            extracted.get("发票编号") or
            ""
        ).strip()
        
        if not invoice_no:
            continue
            
        # Determine personal_dir and filename from doc.filepath
        rel_path = os.path.relpath(doc.filepath, invoice_archive_dir)
        parts = rel_path.split(os.sep)
        if len(parts) >= 2:
            personal_dir = parts[0]
            filename = parts[1]
        else:
            personal_dir = "未分类"
            filename = os.path.basename(doc.filepath)
            
        key = (invoice_no, filename, personal_dir)
        if key not in existing_keys:
            import datetime
            archive_time = doc.upload_time.strftime("%Y-%m-%d %H:%M:%S") if doc.upload_time else datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            synced_records.append({
                "invoice_no": invoice_no,
                "filename": filename,
                "personal_dir": personal_dir,
                "archive_time": archive_time
            })
            existing_keys.add(key)
            has_new = True
            
    # 4. Save back to hidden file if there are any changes
    if len(synced_records) != len(records) or has_new:
        make_file_normal(history_file_path)
        try:
            with open(history_file_path, "w", encoding="utf-8") as f:
                json.dump(synced_records, f, ensure_ascii=False, indent=2)
        finally:
            make_file_hidden(history_file_path)
            
    return synced_records

def check_invoice_duplicate(invoice_archive_dir: str, invoice_no: str, db: Session):
    if not invoice_no:
        return
        
    records = sync_invoice_history_with_db(invoice_archive_dir, db)
    
    duplicate_found = None
    for r in records:
        if r.get("invoice_no") == invoice_no:
            duplicate_found = r
            break
            
    if duplicate_found:
        raise HTTPException(
            status_code=400,
            detail=f"该发票已被归档！发票号码: {invoice_no}，归档文件名: {duplicate_found.get('filename')}，归档人: {duplicate_found.get('personal_dir')}"
        )

def record_invoice_archive(invoice_archive_dir: str, invoice_no: str, filename: str, personal_dir: str, db: Session):
    if not invoice_no:
        return
        
    records = sync_invoice_history_with_db(invoice_archive_dir, db)
    
    import datetime
    new_record = {
        "invoice_no": invoice_no,
        "filename": filename,
        "personal_dir": personal_dir,
        "archive_time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    exists = False
    for r in records:
        if r.get("invoice_no") == invoice_no and r.get("filename") == filename and r.get("personal_dir") == personal_dir:
            exists = True
            break
            
    if not exists:
        records.append(new_record)
        history_file_path = os.path.join(invoice_archive_dir, ".invoice_history.json")
        make_file_normal(history_file_path)
        try:
            with open(history_file_path, "w", encoding="utf-8") as f:
                json.dump(records, f, ensure_ascii=False, indent=2)
        finally:
            make_file_hidden(history_file_path)

class ArchiveDocSchema(BaseModel):
    document_type: str
    extracted_data: str # JSON format string
    summary: str
    link_contract_id: Optional[int] = None
    personal_dir: Optional[str] = None

@app.post("/api/documents/{document_id}/archive", response_model=schemas.Document)
def archive_document(document_id: int, data: ArchiveDocSchema, db: Session = Depends(database.get_db)):
    db_doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if db_doc is None:
        raise HTTPException(status_code=404, detail="Document not found")
        
    try:
        # Update database fields from data payload
        db_doc.document_type = data.document_type
        db_doc.extracted_data = data.extracted_data
        db_doc.summary = data.summary
        db_doc.link_contract_id = data.link_contract_id
        db_doc.is_archived = True
        
        extracted = json.loads(data.extracted_data or "{}")
        
        invoice_no = ""
        archive_dir = ""
        personal_clean = ""
        
        buyer_name = (
            extracted.get("买受方/购买方") or 
            extracted.get("买受方") or 
            extracted.get("购买方") or 
            extracted.get("买方") or 
            ""
        ).strip()
        
        date_val = (
            extracted.get("签订/开票日期") or 
            extracted.get("签订日期") or 
            extracted.get("开票日期") or 
            extracted.get("日期") or 
            ""
        )
        amount_val = extracted.get("价税合计金额") or extracted.get("金额") or ""
        doc_type_val = data.document_type
        
        # Resolve province and county (use pre-resolved values from DB to optimize archiving speed and avoid UI lag)
        prov = db_doc.province
        county = db_doc.county
        if not prov or not county or prov == "未知省份" or county == "未知地区":
            # Fast guess from buyer name without running OCR
            p, c = parse_address_from_text("", buyer_name)
            if p:
                prov = p
            if c:
                county = c
                
            if not prov:
                prov = "未知省份"
            if not county:
                county = "未知地区"
                
        db_doc.province = prov
        db_doc.county = county
        
        # Helper to clean invalid characters for filenames
        def clean_filename(name):
            return re.sub(r'[\\/*?:"<>|]', "", name).strip()
            
        # Clean and map variables
        date_clean = clean_filename(format_date_yyyymmdd(date_val))
        amount_clean = clean_filename(format_amount_clean(amount_val))
        doc_type_clean = clean_filename(get_mapped_doc_type(doc_type_val))
        buyer_clean = clean_filename(buyer_name or "未知买受方")
        
        prov_clean = clean_filename(prov or "未知省份")
        county_clean = clean_filename(county or "未知地区")
        
        config = config_service.get_config()
        if doc_type_clean == "发票" and db_doc.source == "invoice_archive":
            archive_dir = config["invoice_archive_dir"]
            
            invoice_no = (
                extracted.get("合同/发票编号") or
                extracted.get("发票号码") or
                extracted.get("发票编号") or
                ""
            ).strip()
            check_invoice_duplicate(archive_dir, invoice_no, db)
            
            personal_clean = clean_filename(data.personal_dir or "未分类")
            target_dir = os.path.join(archive_dir, personal_clean)
            os.makedirs(target_dir, exist_ok=True)
            
            invoice_content = extracted.get("发票内容") or "商品"
            invoice_content_clean = clean_filename(invoice_content)
            
            filename = os.path.basename(db_doc.filepath)
            _, ext = os.path.splitext(filename)
            ext = ext.lower()
            
            new_filename = f"{date_clean}_{invoice_content_clean}_{amount_clean}{ext}"
            new_filepath = os.path.join(target_dir, new_filename)
            
            counter = 1
            while os.path.exists(new_filepath) and os.path.abspath(db_doc.filepath) != os.path.abspath(new_filepath):
                new_filename = f"{date_clean}_{invoice_content_clean}_{amount_clean}_{counter}{ext}"
                new_filepath = os.path.join(target_dir, new_filename)
                counter += 1
        else:
            archive_dir = config["archive_dir"]
            
            # Map doc types to subfolders: 01合同, 02收货单, 03发票, 04回款
            subfolder_map = {
                "销售合同": "01合同",
                "收货单": "02收货单",
                "发票": "03发票",
                "回款": "04回款"
            }
            subfolder_name = subfolder_map.get(doc_type_clean, f"05{doc_type_clean}")
            
            # Resolve parent folder path
            parent_folder_path = None
            if data.link_contract_id:
                linked_contract = db.query(models.Document).filter(models.Document.id == data.link_contract_id).first()
                if linked_contract and linked_contract.filepath:
                    contract_sub_dir = os.path.dirname(linked_contract.filepath)
                    contract_parent_dir = os.path.dirname(contract_sub_dir)
                    if os.path.exists(contract_parent_dir):
                        parent_folder_path = contract_parent_dir
            
            # Fallback to normal folder structure if not linked or parent dir not found
            if not parent_folder_path:
                prov_dir = os.path.join(archive_dir, prov_clean)
                county_dir = os.path.join(prov_dir, county_clean)
                buyer_dir = os.path.join(county_dir, buyer_clean)
                buyer_folder_name = f"{date_clean}_{doc_type_clean}_{amount_clean}元"
                parent_folder_path = os.path.join(buyer_dir, buyer_folder_name)
            
            target_dir = os.path.join(parent_folder_path, subfolder_name)
            os.makedirs(target_dir, exist_ok=True)
            
            # Construct new filename: [日期]_[买受方]_[类型]_¥[金额]_[状态].[扩展名]
            if doc_type_clean == "销售合同":
                seal_val = extracted.get("盖章状态", "")
                mapped_status = get_mapped_seal_status(seal_val)
            elif doc_type_clean == "收货单":
                mapped_status = extracted.get("收货状态", "未收货")
            elif doc_type_clean == "回款":
                mapped_status = "已回款"
            else:
                mapped_status = "已归档"
            
            filename = os.path.basename(db_doc.filepath)
            _, ext = os.path.splitext(filename)
            ext = ext.lower()
            
            new_filename = f"{date_clean}_{buyer_clean}_{doc_type_clean}_¥{amount_clean}_{mapped_status}{ext}"
            new_filepath = os.path.join(target_dir, new_filename)
            
            counter = 1
            while os.path.exists(new_filepath) and os.path.abspath(db_doc.filepath) != os.path.abspath(new_filepath):
                new_filename = f"{date_clean}_{buyer_clean}_{doc_type_clean}_¥{amount_clean}_{mapped_status}_{counter}{ext}"
                new_filepath = os.path.join(target_dir, new_filename)
                counter += 1
                
        if os.path.abspath(db_doc.filepath) != os.path.abspath(new_filepath):
            shutil.move(db_doc.filepath, new_filepath)
            db_doc.filepath = new_filepath
            db_doc.filename = new_filename
            
        if doc_type_clean == "发票" and db_doc.source == "invoice_archive":
            record_invoice_archive(archive_dir, invoice_no, new_filename, personal_clean, db)
            
    except HTTPException as he:
        raise he
    except Exception as e:
        print("Failed to structure archived document folder:", e)
        
    db.commit()
    db.refresh(db_doc)
    
    # Write directly to the local Excel ledger
    update_excel_ledger_file(db_doc, raise_http=True)
    
    return db_doc


class CropOCRSchema(BaseModel):
    x: float
    y: float
    width: float
    height: float



class CheckDetailsDuplicateSchema(BaseModel):
    doc_id: Optional[int] = None
    buyer: str
    seller: str
    amount: str
    products: str

@app.post("/api/documents/check-details-duplicate")
def api_check_details_duplicate(data: CheckDetailsDuplicateSchema, db: Session = Depends(database.get_db)):
    warning = check_details_duplicate_core(
        doc_id=data.doc_id,
        buyer=data.buyer,
        seller=data.seller,
        amount=data.amount,
        products=data.products,
        db=db
    )
    return {"warning": warning}


@app.get("/api/documents/{document_id}/page-image")
def get_document_page_image(document_id: int, db: Session = Depends(database.get_db)):
    db_doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if db_doc is None or not db_doc.filepath:
        raise HTTPException(status_code=404, detail="Document file not found")
    if not os.path.exists(db_doc.filepath):
        raise HTTPException(status_code=404, detail="Physical file does not exist")
        
    if db_doc.filepath.lower().endswith(".pdf") or "pdf" in db_doc.file_type.lower():
        try:
            with fitz.open(db_doc.filepath) as doc:
                page = doc.load_page(0)
                pix = page.get_pixmap(dpi=150)
                img_data = pix.tobytes("png")
            return StreamingResponse(io.BytesIO(img_data), media_type="image/png")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to render PDF page: {e}")
    else:
        return FileResponse(db_doc.filepath)


@app.post("/api/documents/{document_id}/crop-ocr")
def crop_ocr_document(document_id: int, data: CropOCRSchema, db: Session = Depends(database.get_db)):
    db_doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if db_doc is None or not db_doc.filepath:
        raise HTTPException(status_code=404, detail="Document file not found")
    if not os.path.exists(db_doc.filepath):
        raise HTTPException(status_code=404, detail="Physical file does not exist")
        
    try:
        if db_doc.filepath.lower().endswith(".pdf") or "pdf" in db_doc.file_type.lower():
            with fitz.open(db_doc.filepath) as doc:
                page = doc.load_page(0)
                pix = page.get_pixmap(dpi=150)
                img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
        else:
            img = Image.open(db_doc.filepath)
            
        try:
            orig_w, orig_h = img.size
            
            left = data.x * orig_w
            top = data.y * orig_h
            right = (data.x + data.width) * orig_w
            bottom = (data.y + data.height) * orig_h
            
            left = max(0, min(orig_w - 1, left))
            top = max(0, min(orig_h - 1, top))
            right = max(left + 1, min(orig_w, right))
            bottom = max(top + 1, min(orig_h, bottom))
            
            cropped_img = img.crop((left, top, right, bottom))
            
            text_res = ai_service.extract_text_from_image(cropped_img)
            return {"text": text_res.strip()}
        finally:
            img.close()

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/documents/{document_id}/toggle-receipt", response_model=schemas.Document)
def toggle_document_receipt(document_id: int, db: Session = Depends(database.get_db)):
    db_doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if db_doc is None:
        raise HTTPException(status_code=404, detail="Document not found")
        
    try:
        extracted = json.loads(db_doc.extracted_data or "{}")
    except:
        extracted = {}
        
    current_status = extracted.get("收货状态", "未收货")
    new_status = "已收货" if current_status == "未收货" else "未收货"
    extracted["收货状态"] = new_status
    
    db_doc.extracted_data = json.dumps(extracted, ensure_ascii=False)
    db.commit()
    db.refresh(db_doc)
    return db_doc


@app.get("/api/export")

def export_ledger(db: Session = Depends(database.get_db)):
    documents = db.query(models.Document).filter(
        models.Document.document_type.in_(get_contract_types())
    ).order_by(models.Document.upload_time.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文档台账"
    ws.append(LEDGER_HEADERS)
    
    for doc in documents:
        data_dict = {}
        if doc.extracted_data:
            try:
                data_dict = json.loads(doc.extracted_data)
            except:
                pass
                
        row = [
            doc.id,
            doc.filename,
            doc.document_type,
            data_dict.get("合同/发票编号", ""),
            data_dict.get("出卖方/销售方", ""),
            data_dict.get("买受方/购买方", ""),
            data_dict.get("签订/开票日期", ""),
            data_dict.get("价税合计金额", ""),
            data_dict.get("产品明细", ""),
            data_dict.get("备注", ""),
            "已归档" if doc.is_archived else ("解析完成" if doc.status == "processed" else ("失败" if doc.status == "failed" else "解析中")),
            doc.upload_time.strftime("%Y-%m-%d %H:%M:%S") if doc.upload_time else ""
        ]
        ws.append(row)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ledger.xlsx"}
    )

# Serve frontend static files
frontend_dist = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"))
if os.path.exists(frontend_dist):
    app.mount("/", StaticFiles(directory=frontend_dist, html=True), name="static")

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
